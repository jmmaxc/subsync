import tabulate

import proceso_frases
import os

#
# Esto ya viene en tabulate
#
# from openpyxl import Workbook
import openpyxl


def imprimir(salida):
    print(salida)


def generador_rows_F_Mix(h_sequence, v_sequence, data, data1, hide_zeros=False, nonzero_val=None):
    rows, col_headers = show_F_Mix_interno(h_sequence,
                                           v_sequence,
                                           data,
                                           data1,
                                           hide_zeros,
                                           nonzero_val)
    yield [" "] + col_headers
    for row in rows:
        yield row
    yield [" "]



def show_F_Mix_interno(h_sequence, v_sequence, data, data1, hide_zeros=False, nonzero_val=None):
    rows = []
    col_headers = [c.palabra_str for c in h_sequence]
    row_headers = [c.palabra_str for c in v_sequence]

    pad_headers = data.shape == (len(row_headers) + 1, len(col_headers) + 1)
    if pad_headers:
        row_headers = [" "] + row_headers
        col_headers = [" "] + col_headers
    datos = zip(data, data1)
    for h, d1, d2 in zip(row_headers, data, data1):
        current_row = [h]

        for e, f in zip(d1, d2):
            s = ""
            if e == 0:
                if hide_zeros:
                    s = ''
                else:
                    s = "0"
            else:
                if nonzero_val is not None:
                    s = nonzero_val
                else:
                    s = "{0:2.1f}".format(e)
            if f == 0:
                if hide_zeros:
                    s = "{0}, ".format(s)
                else:
                    s = "{0}, 0".format(s)
            else:
                if nonzero_val is not None:
                    s = "{0}, {1}".format(s, nonzero_val)
                else:
                    s = "{0}, {1}".format(s, f)
            current_row.append(s)
        rows.append(current_row)
    return rows, col_headers


def show_F_Mix(h_sequence, v_sequence, data, data1, hide_zeros=False, nonzero_val=None):
    rows, col_headers = show_F_Mix_interno(h_sequence,
                                           v_sequence,
                                           data,
                                           data1,
                                           hide_zeros,
                                           nonzero_val)
    return tabulate.tabulate(rows, headers=col_headers, tablefmt='fancy_grid')



def generador_rows_F(h_sequence,
                     v_sequence,
                     data,
                     formato_data="{0:2.1f}",
                     hide_zeros=False,
                     nonzero_val=None):
    rows, col_headers = show_F_interno(h_sequence,
                     v_sequence,
                     data,
                     formato_data,
                     hide_zeros,
                     nonzero_val)
    yield [" "] + col_headers
    for row in rows:
        yield row
    yield [" "]

def show_F_interno(h_sequence, v_sequence, data, formato_data="{0:2.1f}", hide_zeros=False, nonzero_val=None):
    rows = []
    col_headers = [c.palabra_str for c in h_sequence]
    row_headers = [c.palabra_str for c in v_sequence]

    pad_headers = data.shape == (len(row_headers) + 1, len(col_headers) + 1)
    if pad_headers:
        row_headers = [" "] + row_headers
        col_headers = [" "] + col_headers
    for h, d in zip(row_headers, data):
        current_row = [h]

        for e in d:
            if e == 0:
                if hide_zeros:
                    current_row.append('')
                else:
                    current_row.append(0)
            else:
                if nonzero_val is not None:
                    current_row.append(nonzero_val)
                else:
                    current_row.append(formato_data.format(e))
        rows.append(current_row)
    return rows, col_headers


def show_F(h_sequence, v_sequence, data, formato_data="{0:2.1f}", hide_zeros=False, nonzero_val=None):
    rows, col_headers = show_F_interno(h_sequence,
                                       v_sequence,
                                       data,
                                       formato_data,
                                       hide_zeros,
                                       nonzero_val)
    return tabulate.tabulate(rows, headers=col_headers, tablefmt='fancy_grid')

def graba_fichero_F (fichero,hoja,h_sequence, v_sequence, data, formato_data="{0:2.1f}", hide_zeros=False, nonzero_val=None):
    graba_campos_fichero_excel(fichero,
                                None,
                                None,
                                hoja,
                                generador_rows_F(h_sequence,
                                                 v_sequence,
                                                 data,
                                                 formato_data,
                                                 hide_zeros,
                                                 nonzero_val
                                ))


def show_alineamiento(titulos, filas):
    rows = []

    for tit, frase in zip(titulos, filas):
        row = []
        row.append(tit)
        for w in frase:

            s = "-"
            if w is not None:
                s = w
                if isinstance(w, proceso_frases.Palabras):
                    s = w.palabra_str
                elif isinstance(w, float):
                    s = "{0:3.2f}".format(w)
            row.append(s)
        rows.append(row)

    return tabulate.tabulate(rows, tablefmt='fancy_grid', )


def imprime(titulo, transcripcion):
    salida = "{:20}".format(titulo)

    fmt = "{1:15}"

    try:
        for w in transcripcion:
            if w is None:
                salida = ("{0}" + fmt).format(salida, "---")
            else:
                if isinstance(w, proceso_frases.Palabras):
                    fmt = "{1:15}"
                    salida = ("{0}" + fmt).format(salida, "{0}".format(w.palabra_str))
                else:
                    if isinstance(w, float):
                        salida = ("{0}" + fmt).format(salida, "{0:3.2f}".format(w))
                    else:
                        salida = ("{0}" + fmt).format(salida, "{0}".format(w))
    except TypeError:
        if isinstance(transcripcion, float):
            salida = ("{0}" + fmt).format(salida, "{0:3.2f}".format(transcripcion))
        else:
            salida = ("{0}" + fmt).format(salida, "{0}".format(transcripcion))

    imprimir(salida)


def graba_estadisticas(nombre_fichero, nombre_hoja, generador_estadisticas):
    s = nombre_fichero.rfind('.')
    fsalida = nombre_fichero + ".xlsx" if s is -1 else nombre_fichero[:s] + ".xlsx"
    if os.path.isfile(fsalida):
        wb = openpyxl.load_workbook(fsalida)
    else:
        wb = openpyxl.Workbook()

    hojas = wb.get_sheet_names()
    if nombre_hoja in hojas:
        std = wb.get_sheet_by_name(nombre_hoja)
        wb.remove_sheet(std)

    ws = wb.create_sheet(nombre_hoja)  #

    for registro in generador_estadisticas:
        ws.append([l for l in registro[0]])
        for fila in registro[1]:
            ws.append(fila)

    wb.save(filename=fsalida)
    wb.close()


def graba_campos_fichero_excel(nombre_fichero, lista_campos, fmt_campos, nombre_hoja, generador_registros):
    s = nombre_fichero.rfind('.')
    fsalida = nombre_fichero + ".xlsx" if s is -1 else nombre_fichero[:s] + ".xlsx"
    if os.path.isfile(fsalida):
        wb = openpyxl.load_workbook(fsalida)
    else:
        wb = openpyxl.Workbook()

    hojas = wb.get_sheet_names()
    if nombre_hoja in hojas:
        std = wb.get_sheet_by_name(nombre_hoja)
        wb.remove_sheet(std)

    ws = wb.create_sheet(nombre_hoja)  #
    if lista_campos is not None:
        ws.append([l for l in lista_campos])
    for fila in generador_registros:
        ws.append(fila)

    wb.save(filename=fsalida)
    wb.close()


def graba_configuracion_fichero_excel(nombre_fichero, lista_configuracion=None, lista_parametros=None):
    s = nombre_fichero.rfind('.')
    fsalida = nombre_fichero + ".xlsx" if s is -1 else nombre_fichero[:s] + ".xlsx"
    wb = openpyxl.load_workbook(fsalida)

    fila = 1
    celdas = []
    if lista_configuracion is not None:
        hojas = wb.get_sheet_names()
        if 'Configuracion' in hojas:
            std = wb.get_sheet_by_name('Configuracion')
            wb.remove_sheet(std)
        ws = wb.create_sheet('Configuracion', 0)

        for configuracion in lista_configuracion:

            fila, columna = imprime_celdas_configuracion(configuracion, fila, 1, 1, celdas)

        for celda in celdas:
            ws.cell(column=celda.columna, row=celda.fila, value=celda.nombre)
            if not celda.categoria:
                ws.cell(column=celda.columna + 1, row=celda.fila, value=celda.valor)

    if lista_parametros is not None:
        hojas = wb.get_sheet_names()
        if 'Parametros' in hojas:
            std = wb.get_sheet_by_name('Parametros')
            wb.remove_sheet(std)
        ws = wb.create_sheet('Parametros')
        cabecera = [w[0] for w in lista_parametros]
        valores = [w[1] for w in lista_parametros]
        ws.append(cabecera)
        ws.append(valores)

    wb.save(filename=fsalida)


def graba_fichero_csv(nombre_fichero, lista_campos, fmt_campos, generador_registros):
    s = nombre_fichero.rfind('.')
    fsalida = nombre_fichero + ".csv" if s is -1 else nombre_fichero[:s] + ".csv"

    with open(fsalida, 'w') as fd:
        salida = ""
        for s in lista_campos:
            salida = salida + s + ","
        salida = salida[0:salida.rfind(",")]  # Quitamos la ultima ,
        fd.write(salida + '\n')

        for linea in generador_registros:
            s1 = zip(fmt_campos, linea)
            salida = ""
            for s in s1:
                salida = salida + s[0].format(s[1]) + ","
            salida = salida[0:salida.rfind(",")]
            fd.write(salida + '\n')


class Celda:
    def __init__(self, fila, columna, categoria, nombre, valor):
        self.fila = fila
        self.columna = columna
        self.categoria = categoria
        self.nombre = nombre
        self.valor = valor

    def informe(self):
        print("Fila: ", self.fila, self.columna,
              "Categoria" if self.categoria else "Elemento", self.nombre, self.valor)


def imprime_celdas_configuracion(lista_configuracion, fila, columna, columna_anterior, celdas):

    i_lista_configuracion = iter(lista_configuracion)
    while True:
        try:
            primero = next(i_lista_configuracion)
            if not isinstance(primero, list) and not isinstance(primero, tuple):
                segundo = next(i_lista_configuracion)
                if not isinstance(segundo, list) and not isinstance(segundo, tuple):
                    # print("Fila: ", fila, "Columna: ", columna, "Nombre: ", primero, " Valor:", segundo)
                    celda = Celda(fila, columna, False, primero, segundo)
                    celdas.append(celda)
                    fila = fila + 1
                else:
                    # print("Fila: ", fila, "Columna: ", columna, "Categoria: ", primero)
                    celda = Celda(fila, columna, True, primero, "")
                    celdas.append(celda)
                    fila = fila + 1
                    fila, columna = imprime_celdas_configuracion(segundo, fila, columna + 1, columna, celdas)

            else:
                fila, columna = imprime_celdas_configuracion(primero, fila, columna, columna, celdas)
        except:
            break
    return fila, columna_anterior
